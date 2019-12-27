from openpyxl import load_workbook


class Excel:
    def __init__(self, file=None, data_only: bool = False, read_only: bool = False, keep_vba: bool = False):
        self.workbook = load_workbook(filename=file, data_only=data_only, read_only=read_only, keep_vba=keep_vba)

    def is_merged_cell(self, row: int, col: int, sheet_name, return_range=False):
        """判断传入的单元格是否为合并单元格

        :param row: 行号
        :param col: 列号
        :param sheet_name: sheet的名字，不传入，则使用 self.excelst
        :param return_range: 是否返回单元格的范围
        :return:
        """
        sheet = self.workbook[sheet_name]
        merged_cells = sheet.merged_cells
        for merged_cell in merged_cells:
            if merged_cell.min_row <= row <= merged_cell.max_row and merged_cell.min_col <= col <= merged_cell.max_col:
                if return_range:
                    return {
                        'min_row': merged_cell.min_row,
                        'max_row': merged_cell.max_row,
                        'min_col': merged_cell.min_col,
                        'max_col': merged_cell.max_col,
                    }
                else:
                    return True
        else:
            return False

    def find_string(self, sheet_name: str, string: [str, int], row=(1, 0), column=(1, 0), full_match: bool = True) -> list:
        """查找指定的文本

        :param sheet_name sheet名
        :param string: 查找的文本
        :param row: 行的搜索范围，二元正整数元组或列表，第一个值表示起始行，默认值为1；第二个值表示结果行，默认值为0，即不指定结束行，以sheet自己的最大行为结束行
        :param column: 列的搜索范围，同row
        :param full_match
        :return: 所有符合条件的单元格的行，列集合的列表， [(row,column), (row,column)]
        """
        for item in [row, column]:
            if not (isinstance(item, tuple) or isinstance(item, list)) or len(item) != 2:
                raise Exception('row or column parameter must be binary tuple')
            for value in item:
                if not isinstance(value, int) or value < 0:
                    raise Exception("binary tuple's value must be greater than or equal to 0")
            if item[1] != 0 and item[1] < item[0]:
                raise Exception('end value must be greater than or equal to start value if end value greater than zero')
        result = []
        ws = self.workbook[sheet_name]
        start_row = row[0] if row[0] > 1 else 1
        end_row = ws.max_row if row[1] == 0 else row[1]
        start_column = column[0] if column[0] > 1 else 1
        end_column = ws.max_column if column[1] == 0 else column[1]

        if start_row > end_row or start_column > end_column:
            raise Exception('end value must be greater than or equal to start value')
        if isinstance(string, int):
            string = str(string)
        for row in range(start_row, end_row + 1):
            for column in range(start_column, end_column + 1):
                if hasattr(ws.cell(row=row, column=column), 'value') is False:
                    continue
                if full_match:
                    if str(ws.cell(row=row, column=column).value).lower() == string.lower():
                        result.append((row, column))
                else:
                    if str(ws.cell(row=row, column=column).value).lower().find(string.lower()) != -1:
                        result.append((row, column))

        return result

    def get_region_values(self, min_row, min_col, max_row=None, max_col=None, sheet_name=None):
        """获取指定区域内的单元格的值，未考虑指定区域存在合并单元格的情况

        :param min_row: 起始行
        :param max_row: 结束行
        :param min_col: 起始列
        :param max_col: 结束列
        :param sheet_name:
        :return:
        """
        sheet = self.workbook[sheet_name]

        max_col = max_col if max_col else sheet.max_column
        max_row = max_row if max_row else sheet.max_row

        values = dict()

        for row in range(min_row, max_row + 1):
            temp = list()
            key = sheet.cell(row=row, column=min_col - 1).value  # min_col的前一列为key值
            for col in range(min_col, max_col + 1):
                cell_value = sheet.cell(row=row, column=col).value
                if cell_value is None:
                    break
                else:
                    temp.append(cell_value)
            values[key] = temp
        return values


